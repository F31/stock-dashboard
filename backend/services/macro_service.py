"""Macro economic data service: US/CN Treasury yields, CPI, PPI, PMI, industrial profit.

Primary sources (EastMoney datacenter — near real-time):
  RPT_ECONOMY_CPI  → NATIONAL_SAME (yoy%), NATIONAL_SEQUENTIAL (mom%)
  RPT_ECONOMY_PPI  → BASE (price index, yoy% = BASE-100), BASE_ACCUMULATE
  RPT_ECONOMY_PMI  → MAKE_INDEX (mfg PMI), NMAKE_INDEX (non-mfg PMI)

Industrial profit: NBS press releases (www.stats.gov.cn/sj/zxfb/)
  Sheet 附表3 of each release's Excel attachment.

Fallback sources (akshare, monthly refresh, data may lag 1-2 months):
  bond_zh_us_rate          → US + CN Treasury yields (always used for yields)
  macro_china_cpi_yearly   → China CPI YoY%
  macro_china_ppi_yearly   → China PPI YoY%
  macro_china_pmi_yearly   → Official NBS manufacturing PMI
  index_pmi_man_cx         → Caixin manufacturing PMI
  macro_china_non_man_pmi  → Official NBS non-manufacturing PMI
"""
import re
import io
import time
import asyncio
import logging
from datetime import datetime, timedelta, date
from typing import Dict, Any, Optional, List

import httpx

logger = logging.getLogger(__name__)

_cache: Dict[str, tuple] = {}
CACHE_TTL_BONDS = 86400   # 24 hours — yields fetched once per day
CACHE_TTL_MACRO = 86400   # 24 hours — monthly releases, once per day

_ak_sem = asyncio.Semaphore(2)

# Per-key locks: prevent thundering-herd when cache is cold and multiple
# clients hit the same endpoint simultaneously.
_fetch_locks: Dict[str, asyncio.Lock] = {}


def _fetch_lock(key: str) -> asyncio.Lock:
    """Return (creating if needed) the asyncio.Lock for a cache key."""
    if key not in _fetch_locks:
        _fetch_locks[key] = asyncio.Lock()
    return _fetch_locks[key]


async def _cached_fetch(cache_key: str, ttl: int, fetch_fn):
    """Double-checked locking wrapper.

    1. L1/L2 cache hit → return immediately (no lock).
    2. Cache miss → acquire per-key lock → re-check (another coroutine may
       have populated it while we waited) → call fetch_fn() → store if truthy.
    Returns the cached / freshly-fetched value, or {} / [] on failure.
    """
    cached = _get_cached(cache_key)
    if cached is not None:
        return cached
    async with _fetch_lock(cache_key):
        cached = _get_cached(cache_key)   # re-check inside lock
        if cached is not None:
            return cached
        result = await fetch_fn()
        if result:                        # don't cache empty/error returns
            _set_cache(cache_key, result, ttl)
        return result

_EM_URL = "https://datacenter-web.eastmoney.com/api/data/v1/get"
_EM_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://data.eastmoney.com/",
}


# ── Two-layer cache: L1 (memory) + L2 (SQLite, survives restarts) ──────────

def _get_cached_db(key: str):
    """L2 cache lookup. On hit, also restores L1 with remaining TTL."""
    try:
        import json
        from database import SessionLocal
        from models import MacroCache
        db = SessionLocal()
        try:
            rec = db.query(MacroCache).filter(MacroCache.cache_key == key).first()
            if rec and rec.expires_at > datetime.utcnow():
                val = json.loads(rec.data_json)
                remaining = (rec.expires_at - datetime.utcnow()).total_seconds()
                _cache[key] = (val, time.time(), remaining)
                logger.debug(f"DB cache hit: {key}, TTL remaining {remaining:.0f}s")
                return val
        finally:
            db.close()
    except Exception as e:
        logger.debug(f"DB cache read error {key}: {e}")
    return None


def _set_cache_db(key: str, val, ttl: int):
    """Write value to L2 DB cache (upsert)."""
    try:
        import json
        from database import SessionLocal
        from models import MacroCache
        db = SessionLocal()
        try:
            expires_at = datetime.utcnow() + timedelta(seconds=ttl)
            data_json = json.dumps(val, ensure_ascii=False, default=str)
            rec = db.query(MacroCache).filter(MacroCache.cache_key == key).first()
            if rec:
                rec.data_json = data_json
                rec.expires_at = expires_at
                rec.updated_at = datetime.utcnow()
            else:
                db.add(MacroCache(cache_key=key, data_json=data_json, expires_at=expires_at))
            db.commit()
        finally:
            db.close()
    except Exception as e:
        logger.debug(f"DB cache write error {key}: {e}")


def _clear_db_cache(key: str):
    try:
        from database import SessionLocal
        from models import MacroCache
        db = SessionLocal()
        try:
            db.query(MacroCache).filter(MacroCache.cache_key == key).delete()
            db.commit()
        finally:
            db.close()
    except Exception as e:
        logger.debug(f"DB cache clear error {key}: {e}")


def _clear_all_db_cache():
    try:
        from database import SessionLocal
        from models import MacroCache
        db = SessionLocal()
        try:
            db.query(MacroCache).delete()
            db.commit()
        finally:
            db.close()
    except Exception as e:
        logger.debug(f"DB cache clear-all error: {e}")


def _get_cached(key: str):
    """Check L1 (memory), then L2 (DB)."""
    if key in _cache:
        val, ts, ttl = _cache[key]
        if time.time() - ts < ttl:
            return val
        del _cache[key]
    return _get_cached_db(key)


def _set_cache(key: str, val, ttl: int):
    """Write to both L1 and L2."""
    _cache[key] = (val, time.time(), ttl)
    _set_cache_db(key, val, ttl)


def clear_macro_cache(keys: list = None) -> None:
    """Clear L1 + L2 cache. keys=None clears all; pass a list for specific keys."""
    if keys is None:
        _cache.clear()
        _clear_all_db_cache()
    else:
        for k in keys:
            _cache.pop(k, None)
            _clear_db_cache(k)


async def _run_ak(fn_name: str, *args, **kwargs):
    import akshare as ak
    fn = getattr(ak, fn_name, None)
    if fn is None:
        raise AttributeError(f"akshare.{fn_name} not found")
    async with _ak_sem:
        return await asyncio.wait_for(
            asyncio.to_thread(fn, *args, **kwargs),
            timeout=25,
        )


async def _fetch_eastmoney(report_name: str, columns: str, page_size: int = 2) -> list:
    """Fetch records from EastMoney datacenter API, sorted by date descending."""
    params = {
        "reportName": report_name,
        "columns": columns,
        "pageSize": page_size,
        "sortColumns": "REPORT_DATE",
        "sortTypes": -1,
        "source": "WEB",
    }
    async with httpx.AsyncClient(timeout=15, headers=_EM_HEADERS) as client:
        resp = await client.get(_EM_URL, params=params)
        resp.raise_for_status()
        body = resp.json()
    if not body.get("success"):
        raise ValueError(f"EastMoney API error: {body.get('message')}")
    return body["result"]["data"] or []


def _fmt_date(d) -> str:
    """Convert datetime.date, datetime string, or string to YYYY-MM-DD."""
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    s = str(d)
    return s[:10]


def _ym(date_str: str) -> str:
    """Trim to YYYY-MM (for period labels)."""
    return str(date_str)[:7]


def _latest_valid(df, value_col: str = "今值"):
    """Return the latest row where value_col is not NaN/None (akshare DataFrames)."""
    import math
    for _, row in df.iloc[::-1].iterrows():
        v = row.get(value_col)
        try:
            if v is not None and not math.isnan(float(v)):
                return row
        except (TypeError, ValueError):
            pass
    return None


# ── Bond Yields (US + CN, akshare — data is up-to-date) ────────────────────

async def fetch_all_yields() -> Dict[str, Any]:
    """
    Fetch US and China Treasury benchmark yields via akshare bond_zh_us_rate.
    Returns:
      {
        "date": "2026-05-15",
        "us": [{"term": "2Y", "value": 4.09}, ...],
        "cn": [{"term": "2Y", "value": 1.27}, ...],
        "spread_10y": -2.82   # CN 10Y minus US 10Y
      }
    """
    async def _fetch():
        try:
            import math
            start = (datetime.now() - timedelta(days=30)).strftime("%Y%m%d")
            df = await _run_ak("bond_zh_us_rate", start_date=start)
            if df is None or df.empty:
                raise ValueError("empty DataFrame")

            def _last_valid_row(anchor_col: str):
                for _, row in df.iloc[::-1].iterrows():
                    try:
                        v = float(row[anchor_col])
                        if not math.isnan(v):
                            return row
                    except (TypeError, ValueError, KeyError):
                        pass
                return None

            us_row = _last_valid_row("美国国债收益率10年")
            cn_row = _last_valid_row("中国国债收益率10年")
            date_str = _fmt_date(df.iloc[-1]["日期"])

            def _get(row, col) -> Optional[float]:
                if row is None:
                    return None
                try:
                    v = float(row[col])
                    return round(v, 4) if not math.isnan(v) else None
                except (TypeError, ValueError, KeyError):
                    return None

            us_yields, cn_yields = [], []
            for term, cn_col, us_col in [
                ("2Y",  "中国国债收益率2年",  "美国国债收益率2年"),
                ("5Y",  "中国国债收益率5年",  "美国国债收益率5年"),
                ("10Y", "中国国债收益率10年", "美国国债收益率10年"),
                ("30Y", "中国国债收益率30年", "美国国债收益率30年"),
            ]:
                cn_v = _get(cn_row, cn_col)
                us_v = _get(us_row, us_col)
                if cn_v is not None:
                    cn_yields.append({"term": term, "value": cn_v})
                if us_v is not None:
                    us_yields.append({"term": term, "value": us_v})

            cn10 = _get(cn_row, "中国国债收益率10年")
            us10 = _get(us_row, "美国国债收益率10年")
            spread = round(cn10 - us10, 4) if (cn10 is not None and us10 is not None) else None
            result = {"date": date_str, "us": us_yields, "cn": cn_yields, "spread_10y": spread}
            logger.info(f"Yields loaded: date={date_str}, US10Y={us10}%, CN10Y={cn10}%")
            return result
        except Exception as e:
            logger.warning(f"fetch_all_yields error: {e}")
            return None

    return await _cached_fetch("macro:yields", CACHE_TTL_BONDS, _fetch) or {}


# ── China CPI (EastMoney primary, akshare fallback) ─────────────────────────

async def fetch_cn_cpi() -> Dict[str, Any]:
    """
    China CPI YoY% and MoM%.
    Returns: {"period": "2026-04", "yoy": 1.2, "mom": 0.3, "prev": 1.0}
    """
    async def _fetch():
        try:
            rows = await _fetch_eastmoney(
                "RPT_ECONOMY_CPI",
                "REPORT_DATE,NATIONAL_SAME,NATIONAL_SEQUENTIAL",
                page_size=2,
            )
            if not rows:
                raise ValueError("empty")
            row = rows[0]
            prev_row = rows[1] if len(rows) >= 2 else None
            result = {
                "period": _ym(row["REPORT_DATE"]),
                "yoy": round(float(row["NATIONAL_SAME"]), 2),
                "mom": round(float(row["NATIONAL_SEQUENTIAL"]), 2) if row.get("NATIONAL_SEQUENTIAL") is not None else None,
                "prev": round(float(prev_row["NATIONAL_SAME"]), 2) if prev_row else None,
            }
            logger.info(f"CPI (EastMoney): {result}")
            return result
        except Exception as e:
            logger.warning(f"fetch_cn_cpi EastMoney error: {e}, falling back to akshare")
            return await _fetch_cn_cpi_akshare() or None

    return await _cached_fetch("macro:cn_cpi", CACHE_TTL_MACRO, _fetch) or {}


async def _fetch_cn_cpi_akshare() -> Dict[str, Any]:
    try:
        df = await _run_ak("macro_china_cpi_yearly")
        if df is None or df.empty:
            raise ValueError("empty")
        row = _latest_valid(df, "今值")
        if row is None:
            raise ValueError("no valid row")
        result = {
            "period": _fmt_date(row["日期"])[:7],
            "yoy": round(float(row["今值"]), 2),
            "mom": None,
            "prev": round(float(row["前值"]), 2) if row.get("前值") == row.get("前值") else None,
        }
        logger.info(f"CPI (akshare fallback): {result}")
        return result
    except Exception as e:
        logger.warning(f"fetch_cn_cpi akshare error: {e}")
        return {}


# ── China PPI (EastMoney primary, akshare fallback) ─────────────────────────

async def fetch_cn_ppi() -> Dict[str, Any]:
    """
    China PPI YoY% (BASE field is price index; yoy% = BASE - 100).
    Returns: {"period": "2026-04", "yoy": 2.8, "prev": 0.5}
    """
    async def _fetch():
        try:
            rows = await _fetch_eastmoney(
                "RPT_ECONOMY_PPI",
                "REPORT_DATE,BASE,BASE_ACCUMULATE",
                page_size=2,
            )
            if not rows:
                raise ValueError("empty")
            row = rows[0]
            prev_row = rows[1] if len(rows) >= 2 else None
            yoy = round(float(row["BASE"]) - 100, 2)
            prev = round(float(prev_row["BASE"]) - 100, 2) if prev_row else None
            result = {"period": _ym(row["REPORT_DATE"]), "yoy": yoy, "prev": prev}
            logger.info(f"PPI (EastMoney): {result}")
            return result
        except Exception as e:
            logger.warning(f"fetch_cn_ppi EastMoney error: {e}, falling back to akshare")
            return await _fetch_cn_ppi_akshare() or None

    return await _cached_fetch("macro:cn_ppi", CACHE_TTL_MACRO, _fetch) or {}


async def _fetch_cn_ppi_akshare() -> Dict[str, Any]:
    try:
        df = await _run_ak("macro_china_ppi_yearly")
        if df is None or df.empty:
            raise ValueError("empty")
        row = _latest_valid(df, "今值")
        if row is None:
            raise ValueError("no valid row")
        result = {
            "period": _fmt_date(row["日期"])[:7],
            "yoy": round(float(row["今值"]), 2),
            "prev": round(float(row["前值"]), 2) if row.get("前值") == row.get("前值") else None,
        }
        logger.info(f"PPI (akshare fallback): {result}")
        return result
    except Exception as e:
        logger.warning(f"fetch_cn_ppi akshare error: {e}")
        return {}


# ── China PMI (EastMoney primary, akshare fallback) ─────────────────────────

async def fetch_cn_pmi() -> Dict[str, Any]:
    """
    China manufacturing + non-manufacturing PMI via EastMoney.
    MAKE_INDEX = manufacturing PMI (制造业)
    NMAKE_INDEX = non-manufacturing PMI (非制造业)

    Returns:
      {
        "mfg_period": "2026-04",
        "mfg_value": 50.3,
        "mfg_prev": 50.4,
        "mfg_source": "官方NBS",
        "svc_period": "2026-04",
        "svc_value": 49.4,
        "svc_prev": 50.1,
      }
    """
    async def _fetch():
        try:
            rows = await _fetch_eastmoney(
                "RPT_ECONOMY_PMI",
                "REPORT_DATE,MAKE_INDEX,NMAKE_INDEX",
                page_size=2,
            )
            if not rows:
                raise ValueError("empty")
            row = rows[0]
            prev_row = rows[1] if len(rows) >= 2 else None
            period = _ym(row["REPORT_DATE"])
            result: Dict[str, Any] = {
                "mfg_period": period,
                "mfg_value":  float(row["MAKE_INDEX"]) if row.get("MAKE_INDEX") is not None else None,
                "mfg_prev":   float(prev_row["MAKE_INDEX"]) if prev_row and prev_row.get("MAKE_INDEX") is not None else None,
                "mfg_source": "官方NBS",
                "svc_period": period,
                "svc_value":  float(row["NMAKE_INDEX"]) if row.get("NMAKE_INDEX") is not None else None,
                "svc_prev":   float(prev_row["NMAKE_INDEX"]) if prev_row and prev_row.get("NMAKE_INDEX") is not None else None,
            }
            logger.info(f"PMI (EastMoney): mfg={result['mfg_value']}, svc={result['svc_value']}")
            return result
        except Exception as e:
            logger.warning(f"fetch_cn_pmi EastMoney error: {e}, falling back to akshare")
            return await _fetch_cn_pmi_akshare() or None

    return await _cached_fetch("macro:cn_pmi", CACHE_TTL_MACRO, _fetch) or {}


async def _fetch_cn_pmi_akshare() -> Dict[str, Any]:
    mfg_official_task = _fetch_pmi_official_mfg()
    svc_task = _fetch_pmi_svc()
    caixin_task = _fetch_pmi_caixin()

    mfg_official, svc, caixin = await asyncio.gather(
        mfg_official_task, svc_task, caixin_task,
        return_exceptions=True,
    )

    mfg_official = mfg_official if not isinstance(mfg_official, Exception) else {}
    svc          = svc          if not isinstance(svc,          Exception) else {}
    caixin       = caixin       if not isinstance(caixin,        Exception) else {}

    if mfg_official.get("value") is not None:
        mfg = mfg_official
        mfg["source"] = "官方NBS"
    elif caixin.get("value") is not None:
        mfg = caixin
        mfg["source"] = "财新"
    else:
        mfg = {}

    result: Dict[str, Any] = {
        "mfg_period": mfg.get("period", ""),
        "mfg_value":  mfg.get("value"),
        "mfg_prev":   mfg.get("prev"),
        "mfg_source": mfg.get("source", ""),
        "svc_period": svc.get("period", ""),
        "svc_value":  svc.get("value"),
        "svc_prev":   svc.get("prev"),
    }
    return result


async def _fetch_pmi_official_mfg() -> Dict[str, Any]:
    try:
        df = await _run_ak("macro_china_pmi_yearly")
        if df is None or df.empty:
            return {}
        row = _latest_valid(df, "今值")
        if row is None:
            return {}
        return {
            "period": _fmt_date(row["日期"])[:7],
            "value":  round(float(row["今值"]), 1),
            "prev":   round(float(row["前值"]), 1) if row.get("前值") == row.get("前值") else None,
        }
    except Exception as e:
        logger.debug(f"Official mfg PMI error: {e}")
        return {}


async def _fetch_pmi_caixin() -> Dict[str, Any]:
    try:
        df = await _run_ak("index_pmi_man_cx")
        if df is None or df.empty:
            return {}
        import math
        valid_rows = df[df["制造业PMI"].apply(
            lambda v: v is not None and not math.isnan(float(v))
        )]
        if valid_rows.empty:
            return {}
        row = valid_rows.iloc[-1]
        prev_row = valid_rows.iloc[-2] if len(valid_rows) >= 2 else None
        return {
            "period": _fmt_date(row["日期"])[:7],
            "value":  round(float(row["制造业PMI"]), 1),
            "prev":   round(float(prev_row["制造业PMI"]), 1) if prev_row is not None else None,
        }
    except Exception as e:
        logger.debug(f"Caixin PMI error: {e}")
        return {}


async def _fetch_pmi_svc() -> Dict[str, Any]:
    try:
        df = await _run_ak("macro_china_non_man_pmi")
        if df is None or df.empty:
            return {}
        row = _latest_valid(df, "今值")
        if row is None:
            return {}
        return {
            "period": _fmt_date(row["日期"])[:7],
            "value":  round(float(row["今值"]), 1),
            "prev":   round(float(row["前值"]), 1) if row.get("前值") == row.get("前值") else None,
        }
    except Exception as e:
        logger.debug(f"Services PMI error: {e}")
        return {}


# ── Industrial Enterprise Profit (NBS press releases) ───────────────────────

_NBS_BASE = "https://www.stats.gov.cn/sj/zxfb/"
_NBS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www.stats.gov.cn/",
}


def _nbs_listing_url(page: int) -> str:
    """NBS listing page URL. Page 1 = index.html, page 2 = index_1.html, etc."""
    if page == 1:
        return _NBS_BASE
    return f"{_NBS_BASE}index_{page - 1}.html"


async def _nbs_find_profit_releases(max_pages: int = 15, count: int = 3) -> list:
    """Scan NBS listing pages to find the most recent N monthly cumulative profit releases."""
    found = []
    seen: set = set()

    async with httpx.AsyncClient(
        timeout=15, verify=False, headers=_NBS_HEADERS, follow_redirects=True
    ) as client:
        for page in range(1, max_pages + 1):
            url = _nbs_listing_url(page)
            try:
                resp = await client.get(url)
                if resp.status_code != 200:
                    continue
                # NBS listing uses title='' attribute for link text (not <a> inner body)
                raw_links = re.findall(
                    r'href="(\./\d{6}/[^"]+\.html)"[^>]*title=[\'"](.*?)[\'"]',
                    resp.text,
                )
                for rel, text in raw_links:
                    # Match monthly cumulative profit releases (skip annual 全年)
                    if not re.search(r'\d{4}年[\d—\-]+月份全国规模以上工业企业利润', text):
                        continue
                    month_dir = re.search(r'\./(\d{6})/', rel)
                    if not month_dir:
                        continue
                    full_url = f"{_NBS_BASE}{month_dir.group(1)}/{rel.split('/')[-1]}"
                    if full_url not in seen:
                        seen.add(full_url)
                        found.append(full_url)
                    if len(found) >= count:
                        return found
            except Exception as e:
                logger.debug(f"NBS listing page {page} error: {e}")

    return found


async def _nbs_parse_profit_release(release_url: str) -> dict:
    """Fetch the NBS release HTML, download its Excel, parse 附表3 for profit data."""
    import openpyxl

    m = re.match(r'(https://www\.stats\.gov\.cn/sj/zxfb/\d{6}/)', release_url)
    if not m:
        raise ValueError(f"Unexpected NBS URL: {release_url}")
    base_dir = m.group(1)

    async with httpx.AsyncClient(
        timeout=30, verify=False, headers=_NBS_HEADERS, follow_redirects=True
    ) as client:
        resp = await client.get(release_url)
        resp.raise_for_status()
        html = resp.text

        # Extract period label (e.g. "2026年1—3月份")
        period = ""
        title_m = re.search(r'<title>([^<]+)</title>', html)
        if title_m:
            pm = re.search(r'(\d{4}年[\d—\-]+月份)', title_m.group(1))
            if pm:
                period = pm.group(1)

        # Find Excel attachment (relative href like ./P020260427...xlsx)
        xls_links = re.findall(r'href="(\./[^"]*\.xlsx?)"', html, re.IGNORECASE)
        if not xls_links:
            raise ValueError(f"No Excel found at {release_url}")

        xls_url = base_dir + xls_links[0].lstrip('./')
        xr = await client.get(xls_url)
        xr.raise_for_status()
        xlsx_bytes = xr.content

    def _parse(data: bytes) -> dict:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if '附表3' not in wb.sheetnames:
            wb.close()
            raise ValueError("Sheet 附表3 not found")

        ws = wb['附表3']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        total_row = elec_row = None
        for row in rows:
            if not row[0]:
                continue
            name = str(row[0]).strip()
            if name == '总计':
                total_row = row
            elif '计算机' in name and '通信' in name:
                elec_row = row

        if total_row is None:
            raise ValueError("总计 row not found in 附表3")

        def _f(v):
            try:
                return round(float(v), 1) if v is not None else None
            except (TypeError, ValueError):
                return None

        def _prev(cur, yoy_pct):
            if cur is None or yoy_pct is None:
                return None
            try:
                return round(cur / (1 + yoy_pct / 100), 1)
            except (TypeError, ZeroDivisionError):
                return None

        total_profit = _f(total_row[5])
        total_yoy    = _f(total_row[6])
        elec_profit  = _f(elec_row[5]) if elec_row else None
        elec_yoy     = _f(elec_row[6]) if elec_row else None

        return {
            "total_profit": total_profit,
            "total_prev":   _prev(total_profit, total_yoy),
            "total_yoy":    total_yoy,
            "elec_profit":  elec_profit,
            "elec_prev":    _prev(elec_profit, elec_yoy),
            "elec_yoy":     elec_yoy,
        }

    parsed = await asyncio.to_thread(_parse, xlsx_bytes)
    parsed["period"] = period
    logger.info(f"NBS profit parsed: period={period}, total={parsed.get('total_profit')}, elec={parsed.get('elec_profit')}")
    return parsed


async def fetch_industrial_profit() -> list:
    """
    Return last 3 monthly cumulative industrial profit periods from NBS.
    Primary: NBS easyquery API (China servers) via the path:
      月度数据 > 工业 > 按行业分工业企业主要经济指标(2018-至今) > 工业企业利润总额
    Fallback: scrape NBS press release Excel attachments (any server, pagination-fixed).
    Each item: {period, total_profit, total_prev, total_yoy, elec_profit, elec_prev, elec_yoy}
    """
    async def _fetch():
        results = await _fetch_profit_from_nbs_easyquery()
        if not results:
            release_urls = await _nbs_find_profit_releases(max_pages=12, count=3)
            logger.info(f"NBS profit releases (press release scraper): {release_urls}")
            for url in release_urls:
                try:
                    data = await _nbs_parse_profit_release(url)
                    if data:
                        results.append(data)
                except Exception as e:
                    logger.warning(f"Failed to parse NBS profit release {url}: {e}")
        return results or None

    return await _cached_fetch("macro:industrial_profit", CACHE_TTL_MACRO, _fetch) or []


# ── NBS Industrial Charts (工业增加值 + 工业出口交货值) ─────────────────────

CACHE_TTL_INDUSTRIAL = 86400  # 24h — data is released monthly


def _nbs_df_to_series(df) -> list:
    """Convert NBS DataFrame (rows=indicators, cols=periods) to chart series list."""
    import re, math
    series = []
    for row_name in df.index:
        points = []
        for col in df.columns:
            m = re.match(r'(\d{4})年(\d{1,2})月', str(col))
            period = f"{m.group(1)}-{m.group(2).zfill(2)}" if m else str(col)[:7]
            try:
                v = float(df.loc[row_name, col])
                v = None if math.isnan(v) else round(v, 2)
            except Exception:
                v = None
            points.append({"period": period, "value": v})
        points.sort(key=lambda x: x["period"])
        points = [p for p in points if p["value"] is not None]
        if points:
            series.append({"name": str(row_name), "data": points})
    return series


def _detect_series_type(name: str) -> str:
    """Classify a series as 'yoy' (percentage) or 'value' (absolute) by its name."""
    if any(kw in name for kw in ("同比", "增长", "增速", "%", "率")):
        return "yoy"
    return "value"


def _annotate_series(series: list) -> list:
    return [{**s, "series_type": _detect_series_type(s.get("name", ""))} for s in series]


def _nbs_period_sort_key(col) -> str:
    """Parse NBS column like '2026年3月' to sortable '2026-03'."""
    m = re.search(r'(\d{4})年(\d{1,2})月', str(col))
    return f"{m.group(1)}-{m.group(2).zfill(2)}" if m else str(col)


def _nbs_period_to_label(col) -> str:
    """'2026年3月' → '2026年1—3月份' (cumulative label used in profit table)."""
    m = re.search(r'(\d{4})年(\d{1,2})月', str(col))
    if not m:
        return str(col)
    year, mo = m.group(1), int(m.group(2))
    return f"{year}年1月份" if mo == 1 else f"{year}年1—{mo}月份"


async def _fetch_profit_from_nbs_easyquery() -> list:
    """
    Fetch 规模以上工业企业利润 via NBS easyquery.
    Data path: 简单查询 > 月度数据 > 工业 > 按行业分工业企业主要经济指标(2018-至今) > 工业企业利润总额
    Returns last 3 periods in the same format as _nbs_parse_profit_release.
    Only works on China-based servers (NBS IP restriction).
    """
    import math

    df = None
    for path in [
        "工业 > 按行业分工业企业主要经济指标(2018-至今) > 工业企业利润总额",
        "工业 > 按行业分工业企业主要经济指标 (2018-至今) > 工业企业利润总额",
        "工业 > 工业企业利润总额",
    ]:
        try:
            df = await _run_ak("macro_china_nbs_nation", kind="月度数据", path=path, period="LAST48")
            if df is not None and not df.empty:
                logger.info(f"NBS profit easyquery OK: path={path!r}, shape={df.shape}")
                break
        except Exception as e:
            logger.debug(f"NBS profit easyquery path={path!r}: {e}")

    if df is None or df.empty:
        return []

    # Sort period columns chronologically
    period_cols = []
    for col in df.columns:
        k = _nbs_period_sort_key(col)
        if re.match(r'\d{4}-\d{2}', k):
            period_cols.append((k, col))
    period_cols = sorted(dict(period_cols).items())  # deduplicate, sort
    if not period_cols:
        return []

    col_by_key = {k: c for k, c in period_cols}

    # Find 总计 and 计算机通信电子 rows
    total_idx = elec_idx = None
    for idx in df.index:
        name = str(idx).strip()
        if total_idx is None and name in ('总计', '合计', '全部工业企业', '全部'):
            total_idx = idx
        if elec_idx is None and '计算机' in name and ('通信' in name or '电子' in name):
            elec_idx = idx

    if total_idx is None:
        logger.warning("NBS profit easyquery: 总计 row not found; available: %s", list(df.index)[:10])
        return []

    def _get(idx, col):
        try:
            f = float(df.loc[idx, col])
            return None if math.isnan(f) else round(f, 1)
        except Exception:
            return None

    def _yoy(cur, prev):
        if cur is None or prev is None or prev == 0:
            return None
        return round((cur - prev) / abs(prev) * 100, 1)

    results = []
    for sort_key, col in period_cols[-3:]:
        y, mo = sort_key.split('-')
        prior_col = col_by_key.get(f"{int(y)-1}-{mo}")

        total = _get(total_idx, col)
        total_prev = _get(total_idx, prior_col) if prior_col else None
        elec = _get(elec_idx, col) if elec_idx else None
        elec_prev = _get(elec_idx, prior_col) if (elec_idx and prior_col) else None

        results.append({
            'period': _nbs_period_to_label(col),
            'total_profit': total,
            'total_prev': total_prev,
            'total_yoy': _yoy(total, total_prev),
            'elec_profit': elec,
            'elec_prev': elec_prev,
            'elec_yoy': _yoy(elec, elec_prev),
        })

    return results


async def _fetch_nbs_indicator(paths: list, period: str = "LAST36", kind: str = "月度数据") -> list:
    """Try each NBS path in order (with given kind); return series list on first success."""
    for path in paths:
        try:
            df = await _run_ak("macro_china_nbs_nation", kind=kind, path=path, period=period)
            if df is not None and not df.empty:
                series = _nbs_df_to_series(df)
                if series:
                    logger.info(f"NBS indicator OK kind={kind!r} path={path!r}, series={[s['name'] for s in series]}")
                    return series
        except Exception as e:
            logger.debug(f"NBS kind={kind!r} path={path!r} error: {e}")
    return []


def _parse_industrial_excel(data: bytes) -> dict:
    """
    Parse NBS monthly industrial production Excel.
    The sheet has 5 columns: [B]=indicator, [C]=当月绝对量, [D]=当月同比%, [E]=累计绝对量, [F]=累计同比%
    Returns current-month 工业增加值 and 出口交货值 data.
    """
    import openpyxl, math

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def _f(v):
        try:
            f = float(v)
            return None if math.isnan(f) else round(f, 1)
        except (TypeError, ValueError):
            return None

    iva_month_yoy = iva_cumul_yoy = None
    exp_month_val = exp_month_yoy = None

    for row in rows:
        if len(row) < 6 or row[1] is None:
            continue
        name = str(row[1]).strip()
        if name == '规模以上工业增加值':
            iva_month_yoy = _f(row[3])   # col D = 当月同比%
            iva_cumul_yoy = _f(row[5])   # col F = 累计同比%
        elif '出口交货值' in name and '亿元' in name:
            exp_month_val = _f(row[2])   # col C = 当月亿元
            exp_month_yoy = _f(row[3])   # col D = 当月同比%

    return {
        'iva_month_yoy': iva_month_yoy,
        'iva_cumul_yoy': iva_cumul_yoy,
        'exp_month_val': exp_month_val,
        'exp_month_yoy': exp_month_yoy,
    }


async def _fetch_nbs_industrial_charts_from_releases(
    max_pages: int = 50, count: int = 24
) -> Dict[str, Any]:
    """
    Build 工业增加值 and 工业出口交货值 time series by scraping NBS monthly press releases.
    Works from any server (not dependent on NBS easyquery IP restriction).
    First run is slow (~10-30s); results cached 24h.
    """
    _sem = asyncio.Semaphore(5)

    async def _get_page(client, page):
        async with _sem:
            url = _nbs_listing_url(page)
            try:
                resp = await client.get(url, timeout=10)
                if resp.status_code != 200:
                    return []
                raw_links = re.findall(
                    r'href="(\./\d{6}/[^"]+\.html)"[^>]*title=[\'"](.*?)[\'"]',
                    resp.text,
                )
                items = []
                for rel, text in raw_links:
                    if not re.search(r'\d{4}年[\d—\-]+月份规模以上工业增加值', text):
                        continue
                    month_dir = re.search(r'\./(\d{6})/', rel)
                    if not month_dir:
                        continue
                    full_url = f"{_NBS_BASE}{month_dir.group(1)}/{rel.split('/')[-1]}"
                    pm = re.search(r'(\d{4})年(?:\d+—)?(\d{1,2})月份', text)
                    if pm:
                        period = f"{pm.group(1)}-{pm.group(2).zfill(2)}"
                        items.append((period, full_url))
                return items
            except Exception as e:
                logger.debug(f"NBS industrial page {page}: {e}")
                return []

    async def _get_excel_data(client, period, release_url):
        async with _sem:
            try:
                m = re.match(r'(https://www\.stats\.gov\.cn/sj/zxfb/\d{6}/)', release_url)
                if not m:
                    return period, None
                base_dir = m.group(1)
                resp = await client.get(release_url, timeout=10)
                resp.raise_for_status()
                xls_links = re.findall(r'href="(\./[^"]*\.xlsx?)"', resp.text, re.IGNORECASE)
                if not xls_links:
                    return period, None
                xls_url = base_dir + xls_links[0].lstrip('./')
                xr = await client.get(xls_url, timeout=20)
                xr.raise_for_status()
                parsed = await asyncio.to_thread(_parse_industrial_excel, xr.content)
                return period, parsed
            except Exception as e:
                logger.debug(f"NBS industrial Excel {period}: {e}")
                return period, None

    async with httpx.AsyncClient(
        timeout=20, verify=False, headers=_NBS_HEADERS, follow_redirects=True
    ) as client:
        # Scan listing pages in batches of 5 until we have enough releases
        seen_periods: set = set()
        releases: list = []

        for batch_start in range(1, max_pages + 1, 5):
            batch_end = min(batch_start + 5, max_pages + 1)
            batch_results = await asyncio.gather(
                *[_get_page(client, p) for p in range(batch_start, batch_end)],
                return_exceptions=True,
            )
            for page_items in batch_results:
                if isinstance(page_items, Exception):
                    continue
                for period, url in page_items:
                    if period not in seen_periods:
                        seen_periods.add(period)
                        releases.append((period, url))
            if len(releases) >= count:
                break

        releases.sort(key=lambda x: x[0], reverse=True)
        releases = releases[:count]

        if not releases:
            logger.info("No NBS industrial production releases found on listing pages")
            return {}

        logger.info(f"NBS industrial: found {len(releases)} releases {[r[0] for r in releases[:5]]}...")

        # Download + parse all Excels concurrently
        excel_results = await asyncio.gather(
            *[_get_excel_data(client, p, u) for p, u in releases],
            return_exceptions=True,
        )

    # Aggregate into time series
    iva_month_pts: List[dict] = []
    iva_cumul_pts: List[dict] = []
    exp_val_pts: List[dict] = []
    exp_yoy_pts: List[dict] = []

    for res in excel_results:
        if isinstance(res, Exception):
            continue
        period, data = res
        if data is None:
            continue
        if data.get('iva_month_yoy') is not None:
            iva_month_pts.append({'period': period, 'value': data['iva_month_yoy']})
        if data.get('iva_cumul_yoy') is not None:
            iva_cumul_pts.append({'period': period, 'value': data['iva_cumul_yoy']})
        if data.get('exp_month_val') is not None:
            exp_val_pts.append({'period': period, 'value': data['exp_month_val']})
        if data.get('exp_month_yoy') is not None:
            exp_yoy_pts.append({'period': period, 'value': data['exp_month_yoy']})

    for pts in [iva_month_pts, iva_cumul_pts, exp_val_pts, exp_yoy_pts]:
        pts.sort(key=lambda x: x['period'])

    out: Dict[str, Any] = {}

    if iva_month_pts or iva_cumul_pts:
        series = []
        if iva_month_pts:
            series.append({'name': '当月同比(%)', 'data': iva_month_pts, 'series_type': 'yoy'})
        if iva_cumul_pts:
            series.append({'name': '累计同比(%)', 'data': iva_cumul_pts, 'series_type': 'yoy'})
        out['industrial_value_added'] = {
            'title': '规模以上工业增加值',
            'unit': '%',
            'chart_type': 'dual_line',
            'series': series,
        }

    if exp_val_pts or exp_yoy_pts:
        series = []
        if exp_val_pts:
            series.append({'name': '当月出口交货值(亿元)', 'data': exp_val_pts, 'series_type': 'value'})
        if exp_yoy_pts:
            series.append({'name': '当月同比(%)', 'data': exp_yoy_pts, 'series_type': 'yoy'})
        out['industrial_export'] = {
            'title': '工业出口交货值',
            'unit': '亿元',
            'chart_type': 'bar_line',
            'series': series,
        }

    return out


async def fetch_industrial_profit_history() -> dict:
    """
    Fetch historical series of industrial profit cumulative YoY rates (~27 months).
    Returns:
      {"total": [{"period": "2024-02", "value": 10.2}, ...],
       "elec":  [{"period": "2024-02", "value": 249.1}, ...],
       "latest_period": "2026-03"}
    """
    async def _fetch():
        result = await _fetch_profit_history_nbs_easyquery()
        if not result:
            result = await _fetch_profit_history_from_releases(count=30)
        return result or None

    return await _cached_fetch("macro:industrial_profit_history", CACHE_TTL_MACRO, _fetch) or {}


async def _fetch_profit_history_nbs_easyquery() -> dict:
    """China-only: compute cumulative YoY series from NBS easyquery absolute values."""
    import math

    df = None
    for path in [
        "工业 > 按行业分工业企业主要经济指标(2018-至今) > 工业企业利润总额",
        "工业 > 按行业分工业企业主要经济指标 (2018-至今) > 工业企业利润总额",
    ]:
        try:
            df = await _run_ak("macro_china_nbs_nation", kind="月度数据", path=path, period="LAST48")
            if df is not None and not df.empty:
                break
        except Exception as e:
            logger.debug(f"NBS profit history easyquery path={path!r}: {e}")

    if df is None or df.empty:
        return {}

    period_cols = []
    for col in df.columns:
        k = _nbs_period_sort_key(col)
        if re.match(r'\d{4}-\d{2}', k):
            period_cols.append((k, col))
    period_cols = sorted(dict(period_cols).items())
    if not period_cols:
        return {}

    col_by_key = {k: c for k, c in period_cols}

    total_idx = elec_idx = None
    for idx in df.index:
        name = str(idx).strip()
        if total_idx is None and name in ('总计', '合计', '全部工业企业', '全部'):
            total_idx = idx
        if elec_idx is None and '计算机' in name and ('通信' in name or '电子' in name):
            elec_idx = idx

    if total_idx is None:
        return {}

    def _get(idx, col):
        try:
            f = float(df.loc[idx, col])
            return None if math.isnan(f) else round(f, 1)
        except Exception:
            return None

    total_pts: List[dict] = []
    elec_pts: List[dict] = []

    for sort_key, col in period_cols:
        y, mo = sort_key.split('-')
        prior_col = col_by_key.get(f"{int(y)-1}-{mo}")
        if prior_col is None:
            continue
        total = _get(total_idx, col)
        total_prev = _get(total_idx, prior_col)
        if total is not None and total_prev is not None and total_prev != 0:
            yoy = round((total - total_prev) / abs(total_prev) * 100, 1)
            total_pts.append({"period": sort_key, "value": yoy})
        if elec_idx:
            elec = _get(elec_idx, col)
            elec_prev = _get(elec_idx, prior_col)
            if elec is not None and elec_prev is not None and elec_prev != 0:
                yoy = round((elec - elec_prev) / abs(elec_prev) * 100, 1)
                elec_pts.append({"period": sort_key, "value": yoy})

    if not total_pts:
        return {}

    latest = max(p["period"] for p in total_pts)
    logger.info(f"NBS profit history easyquery: {len(total_pts)} periods, latest={latest}")
    return {"total": total_pts, "elec": elec_pts, "latest_period": latest}


async def _fetch_profit_history_from_releases(count: int = 30) -> dict:
    """Batch-scrape NBS profit press releases to build historical YoY series."""
    release_urls = await _nbs_find_profit_releases(max_pages=25, count=count)
    if not release_urls:
        return {}

    sem = asyncio.Semaphore(5)

    async def _parse_one(url):
        async with sem:
            try:
                return await _nbs_parse_profit_release(url)
            except Exception as e:
                logger.debug(f"profit history release {url}: {e}")
                return None

    raw = await asyncio.gather(*[_parse_one(u) for u in release_urls], return_exceptions=True)

    total_pts: List[dict] = []
    elec_pts: List[dict] = []

    for res in raw:
        if isinstance(res, Exception) or res is None:
            continue
        period_label = res.get("period", "")
        pm = re.search(r'(\d{4})年(?:\d+[—\-])?(\d{1,2})月份', period_label)
        if not pm:
            continue
        period_key = f"{pm.group(1)}-{pm.group(2).zfill(2)}"
        if res.get("total_yoy") is not None:
            total_pts.append({"period": period_key, "value": res["total_yoy"]})
        if res.get("elec_yoy") is not None:
            elec_pts.append({"period": period_key, "value": res["elec_yoy"]})

    total_pts.sort(key=lambda x: x["period"])
    elec_pts.sort(key=lambda x: x["period"])

    if not total_pts:
        return {}

    latest = max(p["period"] for p in total_pts)
    logger.info(f"NBS profit history scraper: {len(total_pts)} periods, latest={latest}")
    return {"total": total_pts, "elec": elec_pts, "latest_period": latest}


async def fetch_nbs_industrial_charts() -> Dict[str, Any]:
    """
    Fetch 工业增加值 and 工业出口交货值 time series.

    Data paths (from NBS website):
      工业增加值:    经济图表 > 月度经济图表 > 工业生产 > 工业增加值
      工业出口交货值: 经济图表 > 月度经济图表 > 工业生产 > 工业出口交货值

    Priority:
      1. NBS easyquery via akshare — "月度经济图表" kind (user-specified paths, China servers)
      2. NBS easyquery via akshare — "月度数据" kind (alternate paths, China servers)
      3. NBS press release Excel scraper — any server, builds 24-month series
      4. Jin10 fallback — any server, 工业增加值 YoY only
    """
    async def _fetch():
        result: Dict[str, Any] = {}

        iva_series = await _fetch_nbs_indicator(["工业生产 > 工业增加值"], kind="月度经济图表")
        if not iva_series:
            iva_series = await _fetch_nbs_indicator([
                "工业 > 规模以上工业增加值",
                "工业 > 规模以上工业 > 工业增加值",
            ])

        exp_series = await _fetch_nbs_indicator(["工业生产 > 工业出口交货值"], kind="月度经济图表")
        if not exp_series:
            exp_series = await _fetch_nbs_indicator([
                "工业 > 工业出口交货值",
                "工业 > 工业总产值及出口交货值 > 工业出口交货值",
                "工业 > 工业分大类行业出口交货值(2018-至今) > 总计",
            ])

        if iva_series:
            result["industrial_value_added"] = {
                "title": "规模以上工业增加值", "unit": "%",
                "chart_type": "dual_line", "series": _annotate_series(iva_series),
            }
        if exp_series:
            result["industrial_export"] = {
                "title": "工业出口交货值", "unit": "亿元",
                "chart_type": "bar_line", "series": _annotate_series(exp_series),
            }

        need_iva = not result.get('industrial_value_added')
        need_exp = not result.get('industrial_export')
        if need_iva or need_exp:
            try:
                scraper = await _fetch_nbs_industrial_charts_from_releases(max_pages=50, count=24)
                if need_iva and scraper.get('industrial_value_added'):
                    result['industrial_value_added'] = scraper['industrial_value_added']
                if need_exp and scraper.get('industrial_export'):
                    result['industrial_export'] = scraper['industrial_export']
            except Exception as e:
                logger.warning(f"NBS industrial press release scraper failed: {e}")

        if not result.get('industrial_value_added'):
            try:
                df = await _run_ak("macro_china_industrial_production_yoy")
                if df is not None and not df.empty:
                    import math
                    points = []
                    for _, row in df.iterrows():
                        try:
                            v = float(row.get("今值") or row.get("当前值") or 0)
                            period = str(row.get("日期", ""))[:7]
                            if v == v and period:
                                points.append({"period": period, "value": round(v, 2)})
                        except Exception:
                            pass
                    points.sort(key=lambda x: x["period"])
                    points = [p for p in points if p["value"] is not None][-36:]
                    if points:
                        result["industrial_value_added"] = {
                            "title": "规模以上工业增加值", "unit": "%",
                            "chart_type": "dual_line",
                            "series": [{"name": "当月同比(%)", "data": points, "series_type": "yoy"}],
                        }
            except Exception as e:
                logger.debug(f"Jin10 工业增加值 fallback error: {e}")

        logger.info(f"NBS industrial charts fetched: keys={list(result.keys())}")
        return result or None

    return await _cached_fetch("macro:nbs_industrial_charts", CACHE_TTL_INDUSTRIAL, _fetch) or {}


# ── US Macro (FRED API) ────────────────────────────────────────────────────

async def fetch_us_fred_macro() -> Dict[str, Any]:
    """
    从 FRED API 获取美国宏观指标（复用 premarket.collector 的同步函数）。
    返回格式: {key: {label, period, value}, ...}
    缓存 12 小时（同 CN 月度数据）。
    """
    async def _fetch():
        try:
            from premarket.collector import _fetch_fred_us_macro
            import os
            api_key = os.environ.get("FRED_API_KEY", "").strip()
            if not api_key:
                return None
            data, errors = await asyncio.to_thread(_fetch_fred_us_macro, api_key)
            if errors:
                logger.warning("FRED fetch partial errors: %s", errors)
            return data or None
        except Exception as e:
            logger.warning("fetch_us_fred_macro error: %s", e)
            return None

    return await _cached_fetch("macro:us_fred", CACHE_TTL_MACRO, _fetch) or {}


# ── Aggregate ───────────────────────────────────────────────────────────────

async def get_macro_data() -> Dict[str, Any]:
    """Fetch all macro indicators concurrently."""
    yields, cpi, ppi, pmi, us_fred = await asyncio.gather(
        fetch_all_yields(),
        fetch_cn_cpi(),
        fetch_cn_ppi(),
        fetch_cn_pmi(),
        fetch_us_fred_macro(),
        return_exceptions=True,
    )

    def _safe(v):
        return {} if isinstance(v, Exception) else (v or {})

    return {
        "yields":   _safe(yields),
        "cn_cpi":   _safe(cpi),
        "cn_ppi":   _safe(ppi),
        "cn_pmi":   _safe(pmi),
        "us_fred":  _safe(us_fred),
    }
